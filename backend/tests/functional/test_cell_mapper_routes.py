"""Cell mapper route functional tests."""
import io
import os
import tempfile
import openpyxl

from app.models.servicer import Servicer
from app.models.deal import Deal
from app.models.variable import VariableDefinition
from app.models.variable_mapping import VariableMapping
from app.models.processing import ProcessingRun, ExtractedValue
from decimal import Decimal


def _make_tape_bytes(sheets: dict[str, list[list]]) -> bytes:
    """Create an xlsx in memory and return bytes."""
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _make_tape_file(sheets: dict[str, list[list]]) -> str:
    """Create an xlsx on disk and return path."""
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    path = tempfile.mktemp(suffix=".xlsx")
    wb.save(path)
    return path


def _setup(db):
    s = db.query(Servicer).first()
    if not s:
        s = Servicer(name="WF", short_code="WF")
        db.add(s)
        db.flush()
    d = Deal(name="CELL-TEST", servicer_id=s.id, created_by="testuser")
    db.add(d)
    db.flush()
    return d


def test_tape_grid_all_sheets(admin_client, db):
    """GET /tape-grid returns all sheets when no sheet param."""
    deal = _setup(db)
    # Create a run with a tape file
    tape_path = _make_tape_file({
        "Summary": [["Name", "Value"], ["Total", 1000]],
        "Details": [["A", "B", "C"]],
    })
    try:
        run = ProcessingRun(
            deal_id=deal.id, report_period="2026-04", created_by="testuser",
            status="extracted", tape_file_path=tape_path,
        )
        db.add(run)
        db.flush()

        resp = admin_client.get(f"/api/deals/{deal.id}/tape-grid")
        assert resp.status_code == 200
        data = resp.json()
        assert data["sheet_names"] == ["Summary", "Details"]
        assert len(data["sheets"]) == 2
        assert data["sheets"][0]["sheet_name"] == "Summary"
        assert data["sheets"][0]["rows"][0]["cells"][0] == "Name"
        assert data["sheets"][0]["rows"][1]["cells"][1] == 1000
    finally:
        os.unlink(tape_path)


def test_tape_grid_single_sheet(admin_client, db):
    """GET /tape-grid?sheet=Summary returns only that sheet."""
    deal = _setup(db)
    tape_path = _make_tape_file({
        "Summary": [["X", 42]],
        "Other": [["Y"]],
    })
    try:
        run = ProcessingRun(
            deal_id=deal.id, report_period="2026-04", created_by="testuser",
            status="extracted", tape_file_path=tape_path,
        )
        db.add(run)
        db.flush()

        resp = admin_client.get(f"/api/deals/{deal.id}/tape-grid?sheet=Summary")
        assert resp.status_code == 200
        data = resp.json()
        assert "sheet" in data
        assert data["sheet"]["sheet_name"] == "Summary"
        assert "sheets" not in data
    finally:
        os.unlink(tape_path)


def test_tape_grid_no_tape(admin_client, db):
    """GET /tape-grid returns 404 when no tape exists."""
    deal = _setup(db)
    resp = admin_client.get(f"/api/deals/{deal.id}/tape-grid")
    assert resp.status_code == 404


def test_reextract_variable_endpoint(admin_client, db):
    """POST /reextract-variable updates a single extracted value."""
    deal = _setup(db)
    v = VariableDefinition(name="svc_fee_tape", scope="system", data_type="decimal")
    db.add(v)
    db.flush()

    tape_path = _make_tape_file({"Fees": [
        [None] * 3,  # row 1
        [None] * 3,  # row 2
        [None, None, 11303.35],  # row 3 → C3
    ]})
    try:
        run = ProcessingRun(
            deal_id=deal.id, report_period="2026-04", created_by="testuser",
            status="extracted", tape_file_path=tape_path,
        )
        db.add(run)
        db.flush()

        db.add(VariableMapping(
            deal_id=deal.id, variable_id=v.id,
            sheet_name="Fees", column_letter="C", row_number=3,
        ))
        # Seed old extracted value
        db.add(ExtractedValue(
            run_id=run.id, variable_id=v.id, variable_name="svc_fee_tape",
            sheet_name="Fees", cell_ref="C1", raw_value="999",
            parsed_value=Decimal("999"), data_type="decimal",
        ))
        db.flush()

        resp = admin_client.post(f"/api/deals/{deal.id}/runs/{run.id}/reextract-variable/{v.id}")
        assert resp.status_code == 200
        data = resp.json()
        assert data["variable"] == "svc_fee_tape"
        assert data["cell"] == "C3"
        assert float(data["parsed"]) == 11303.35
    finally:
        os.unlink(tape_path)


def test_reextract_variable_no_mapping(admin_client, db):
    """POST /reextract-variable returns 400 when no mapping exists."""
    deal = _setup(db)
    v = VariableDefinition(name="unknown_var", scope="system", data_type="decimal")
    db.add(v)
    db.flush()

    tape_path = _make_tape_file({"Sheet1": [[1]]})
    try:
        run = ProcessingRun(
            deal_id=deal.id, report_period="2026-04", created_by="testuser",
            status="extracted", tape_file_path=tape_path,
        )
        db.add(run)
        db.flush()

        resp = admin_client.post(f"/api/deals/{deal.id}/runs/{run.id}/reextract-variable/{v.id}")
        assert resp.status_code == 400
        assert "No mapping" in resp.json()["detail"]
    finally:
        os.unlink(tape_path)
