"""Reextract variable unit tests."""
import os
import tempfile
from decimal import Decimal
import openpyxl

from app.models.servicer import Servicer
from app.models.deal import Deal
from app.models.variable import VariableDefinition
from app.models.variable_mapping import VariableMapping
from app.models.processing import ProcessingRun, ExtractedValue
from app.processing.service import ProcessingService


def _make_tape(sheets: dict[str, dict[str, object]]) -> str:
    """Create a temp xlsx. sheets = {"Sheet1": {"B2": 100.50, "C7": 500}}"""
    wb = openpyxl.Workbook()
    first = True
    for name, cells in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for ref, val in cells.items():
            ws[ref] = val
    path = tempfile.mktemp(suffix=".xlsx")
    wb.save(path)
    return path


def _setup(db):
    s = Servicer(name="WF", short_code="WF")
    db.add(s)
    db.flush()
    deal = Deal(name="RE-TEST", servicer_id=s.id, created_by="t")
    db.add(deal)
    db.flush()
    v = VariableDefinition(name="svc_fee", scope="system", data_type="decimal")
    db.add(v)
    db.flush()
    return deal, v


def test_reextract_success(db):
    deal, var = _setup(db)

    tape_path = _make_tape({"Fees": {"C13": 11303.35}})
    try:
        run = ProcessingRun(deal_id=deal.id, report_period="2026-04", created_by="t", status="extracted", tape_file_path=tape_path)
        db.add(run)
        db.flush()

        db.add(VariableMapping(deal_id=deal.id, variable_id=var.id, sheet_name="Fees", column_letter="C", row_number=13))
        db.flush()

        # Add existing extracted value (old cell)
        db.add(ExtractedValue(run_id=run.id, variable_id=var.id, variable_name="svc_fee",
                              sheet_name="Fees", cell_ref="C7", raw_value="500", parsed_value=Decimal("500"),
                              data_type="decimal"))
        db.flush()

        svc = ProcessingService(db)
        result = svc.reextract_variable(run.id, var.id)

        assert result.cell_ref == "C13"
        assert result.parsed_value == Decimal("11303.35")
        assert result.sheet_name == "Fees"
        assert result.warning is None
    finally:
        os.unlink(tape_path)


def test_reextract_missing_mapping(db):
    deal, var = _setup(db)
    tape_path = _make_tape({"Sheet1": {"A1": 1}})
    try:
        run = ProcessingRun(deal_id=deal.id, report_period="2026-04", created_by="t", status="extracted", tape_file_path=tape_path)
        db.add(run)
        db.flush()
        # No mapping added

        svc = ProcessingService(db)
        import pytest
        with pytest.raises(ValueError, match="No mapping"):
            svc.reextract_variable(run.id, var.id)
    finally:
        os.unlink(tape_path)


def test_reextract_empty_cell(db):
    deal, var = _setup(db)
    tape_path = _make_tape({"Fees": {}})  # C13 is empty
    try:
        run = ProcessingRun(deal_id=deal.id, report_period="2026-04", created_by="t", status="extracted", tape_file_path=tape_path)
        db.add(run)
        db.flush()

        db.add(VariableMapping(deal_id=deal.id, variable_id=var.id, sheet_name="Fees", column_letter="C", row_number=13))
        db.flush()

        svc = ProcessingService(db)
        result = svc.reextract_variable(run.id, var.id)

        assert result.parsed_value is None
        assert result.warning is not None
        assert "empty" in result.warning.lower()
    finally:
        os.unlink(tape_path)
