"""ExcelReader grid mode unit tests."""

import os
import tempfile
import openpyxl
from app.utils.excel_reader import ExcelReader


def _make_workbook(sheets: dict[str, list[list]]) -> str:
    """Create a temp xlsx with specified sheets and data."""
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    path = tempfile.mktemp(suffix=".xlsx")
    wb.save(path)
    return path


def test_read_sheet_grid_normal():
    path = _make_workbook(
        {
            "Summary": [
                ["Deal Name", "AMORT 2024-1", None],
                ["Report Date", "04/01/2026", None],
                ["Total Collections", 4521338.42, None],
            ],
        }
    )
    try:
        with ExcelReader(path) as reader:
            grid = reader.read_sheet_grid("Summary")
            assert grid["sheet_name"] == "Summary"
            assert "A" in grid["column_letters"]
            assert "B" in grid["column_letters"]
            assert len(grid["rows"]) == 3
            assert grid["rows"][0]["row_number"] == 1
            assert grid["rows"][0]["cells"][0] == "Deal Name"
            assert grid["rows"][2]["cells"][1] == 4521338.42
    finally:
        os.unlink(path)


def test_read_sheet_grid_respects_limits():
    """Grid should cap at max_rows and max_cols."""
    path = _make_workbook(
        {
            "Big": [[f"r{r}c{c}" for c in range(30)] for r in range(150)],
        }
    )
    try:
        with ExcelReader(path) as reader:
            grid = reader.read_sheet_grid("Big", max_rows=10, max_cols=5)
            assert len(grid["rows"]) == 10
            assert len(grid["column_letters"]) == 5
            assert grid["column_letters"] == ["A", "B", "C", "D", "E"]
    finally:
        os.unlink(path)


def test_col_num_to_letter():
    assert ExcelReader._col_num_to_letter(1) == "A"
    assert ExcelReader._col_num_to_letter(26) == "Z"
    assert ExcelReader._col_num_to_letter(27) == "AA"
    assert ExcelReader._col_num_to_letter(52) == "AZ"


def test_read_sheet_grid_multiple_sheets():
    path = _make_workbook(
        {
            "Sheet1": [[1, 2], [3, 4]],
            "Sheet2": [["a", "b"]],
        }
    )
    try:
        with ExcelReader(path) as reader:
            names = reader.get_sheet_names()
            assert "Sheet1" in names
            assert "Sheet2" in names

            g1 = reader.read_sheet_grid("Sheet1")
            assert len(g1["rows"]) == 2
            g2 = reader.read_sheet_grid("Sheet2")
            assert len(g2["rows"]) == 1
            assert g2["rows"][0]["cells"][0] == "a"
    finally:
        os.unlink(path)


def test_context_manager():
    path = _make_workbook({"S": [[1]]})
    try:
        with ExcelReader(path) as reader:
            val = reader.get_cell_value("S", "A", 1)
            assert val == 1
    finally:
        os.unlink(path)
