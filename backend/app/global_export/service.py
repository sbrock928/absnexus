"""Global export service — templates, row configs, CSV generation."""

import csv
import hashlib
import io
import os
from decimal import Decimal

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.core import settings
from app.formulas.engine import FormulaEngine
from app.global_export.dao import GlobalExportDAO
from app.models.dag import DagNode
from app.models.deal import Deal
from app.models.global_export import GlobalExportColumn, DealExportRow, DealExportCell
from app.models.processing import ProcessingRun, ExtractedValue, ExecutionStep
from app.schemas.global_export import (
    DealExportConfigResponse,
    DealExportRowResponse,
    DealExportCellResponse,
    GlobalTemplateResponse,
    TemplateWithColumnsResponse,
    GlobalColumnResponse,
)
from app.tranches.dao import TrancheDAO


class GlobalExportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.dao = GlobalExportDAO(db)
        self.tranche_dao = TrancheDAO(db)
        self.formula_engine = FormulaEngine()

    # ── Templates ──

    def list_templates(self) -> list[GlobalTemplateResponse]:
        templates = self.dao.list_templates()
        return [
            GlobalTemplateResponse(id=t.id, name=t.name, description=t.description)
            for t in templates
        ]

    def get_template_with_columns(self, template_id: int) -> TemplateWithColumnsResponse:
        t = self.dao.get_template(template_id)
        if t is None:
            raise ValueError("Template not found.")
        cols = self.dao.list_columns(template_id)
        return TemplateWithColumnsResponse(
            template=GlobalTemplateResponse(id=t.id, name=t.name, description=t.description),
            columns=[GlobalColumnResponse.model_validate(c) for c in cols],
        )

    # ── Deal export config ──

    def get_deal_config(self, deal_id: int, template_id: int) -> DealExportConfigResponse:
        """Get full export row config for a deal+template."""
        rows = self.dao.list_deal_rows(deal_id, template_id)
        row_responses = []
        for row in rows:
            cells = self.dao.list_cells_for_row(row.id)
            node = self.db.query(DagNode).filter(DagNode.id == row.node_id).first()
            row_responses.append(
                DealExportRowResponse(
                    id=row.id,
                    node_id=row.node_id,
                    node_key=node.key if node else None,
                    node_name=node.name if node else None,
                    row_order=row.row_order,
                    identifier_group=row.identifier_group,
                    cells=[
                        DealExportCellResponse(
                            id=c.id,
                            column_id=c.column_id,
                            value_source=c.value_source,
                            source_ref=c.source_ref,
                        )
                        for c in cells
                    ],
                )
            )
        return DealExportConfigResponse(rows=row_responses)

    def save_deal_config(
        self,
        deal_id: int,
        template_id: int,
        rows_data: list[dict],
    ) -> DealExportConfigResponse:
        self.dao.save_deal_config(deal_id, template_id, rows_data)
        return self.get_deal_config(deal_id, template_id)

    # ── CSV generation ──

    def generate_csv(self, run: ProcessingRun, template_id: int) -> tuple[str, str]:
        """Generate CSV for a run using global template + deal row config."""
        if run.status not in ("executed", "completed"):
            raise ValueError(f"Cannot export: run status is '{run.status}'.")

        cols = self.dao.list_columns(template_id)
        if not cols:
            raise ValueError("No columns configured for this template.")

        context = self._build_context(run)
        content = self._generate(run, cols, template_id, context)

        # Save file
        deal = self.db.query(Deal).filter(Deal.id == run.deal_id).first()
        template = self.dao.get_template(template_id)
        deal_name = deal.name.replace(" ", "_") if deal else f"deal_{run.deal_id}"
        tmpl_name = template.name.replace(" ", "_") if template else f"tmpl_{template_id}"
        period = run.report_period or "unknown"
        filename = f"{deal_name}_{tmpl_name}_{period.replace('-', '')}.csv"
        base_dir = (deal.export_directory_override if deal else None) or settings.export_directory
        dest_dir = os.path.join(base_dir, str(run.deal_id), period)
        os.makedirs(dest_dir, exist_ok=True)
        file_path = os.path.join(dest_dir, filename)

        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)

        file_hash = hashlib.sha256(content.encode()).hexdigest()
        run.export_file_path = file_path
        run.export_file_hash = file_hash
        self.db.flush()

        return file_path, file_hash

    def preview(self, deal_id: int, template_id: int, run_id: int | None = None) -> str:
        """Generate preview CSV text without saving to disk."""
        cols = self.dao.list_columns(template_id)
        if not cols:
            return ""

        if run_id:
            run = self.db.query(ProcessingRun).filter(ProcessingRun.id == run_id).first()
            if run:
                context = self._build_context(run)
                return self._generate(run, cols, template_id, context)

        # Placeholder rows from deal's configured export rows
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow([c.header_label for c in cols])
        for row_values in self._placeholder_rows(deal_id, template_id, cols):
            writer.writerow(row_values)
        return out.getvalue()

    def preview_structured(self, deal_id: int, template_id: int) -> dict:
        """Return JSON-shaped preview: {columns: [...], rows: [[...], ...]}."""
        cols = self.dao.list_columns(template_id)
        column_headers = [c.header_label for c in cols]
        rows = self._placeholder_rows(deal_id, template_id, cols)
        return {"columns": column_headers, "rows": rows}

    def preview_xlsx(self, deal_id: int, template_id: int) -> bytes:
        """Build an .xlsx workbook with placeholder values for the template."""
        cols = self.dao.list_columns(template_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Preview"

        for idx, col in enumerate(cols, start=1):
            ws.cell(row=1, column=idx, value=col.header_label)

        for row_idx, row_values in enumerate(
            self._placeholder_rows(deal_id, template_id, cols), start=2
        ):
            for col_idx, value in enumerate(row_values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _placeholder_rows(
        self,
        deal_id: int,
        template_id: int,
        cols: list[GlobalExportColumn],
    ) -> list[list[str]]:
        """Build placeholder rows mirroring the deal's configured DealExportRows.

        Each cell shows `<source_ref>` or `<meta_field>` so users can see what
        will populate it at run time.
        """
        deal_rows = self.dao.list_deal_rows(deal_id, template_id)
        # Fall back to a single placeholder row when no deal config exists yet.
        if not deal_rows:
            return [[self._placeholder(c) for c in cols]]

        rendered: list[list[str]] = []
        for row in deal_rows:
            cells = self.dao.list_cells_for_row(row.id)
            cells_by_col = {c.column_id: c for c in cells}
            values: list[str] = []
            for col in cols:
                cell = cells_by_col.get(col.id)
                if cell is None:
                    values.append(self._placeholder(col))
                    continue
                if cell.value_source == "literal":
                    values.append(cell.source_ref or "")
                else:
                    # node/variable/formula/run_meta/deal_meta — show the ref as a placeholder token
                    values.append(
                        f"<{cell.source_ref}>" if cell.source_ref else self._placeholder(col)
                    )
            rendered.append(values)
        return rendered

    def _build_context(self, run: ProcessingRun) -> dict[str, Decimal]:
        """Build full formula context from tape values + execution results."""
        context: dict[str, Decimal] = {}

        # Tape values
        extracted = self.db.query(ExtractedValue).filter(ExtractedValue.run_id == run.id).all()
        for ev in extracted:
            if ev.parsed_value is not None:
                context[ev.variable_name] = ev.parsed_value

        # Execution step results (by node key)
        steps = self.db.query(ExecutionStep).filter(ExecutionStep.run_id == run.id).all()
        for s in steps:
            if s.result is not None:
                context[s.node_key] = s.result

        # Tranche context
        from app.tranches.service import TrancheService

        tranche_ctx = TrancheService(self.db).build_tranche_context(
            run.deal_id,
            run.report_period or "",
        )
        context.update(tranche_ctx)

        return context

    def _generate(
        self,
        run: ProcessingRun,
        cols: list[GlobalExportColumn],
        template_id: int,
        context: dict[str, Decimal],
    ) -> str:
        deal = self.db.query(Deal).filter(Deal.id == run.deal_id).first()

        # Load deal export rows grouped by node
        deal_rows = self.dao.list_deal_rows(run.deal_id, template_id)

        # Build cells lookup: row_id → {column_id: cell}
        cells_by_row: dict[int, dict[int, DealExportCell]] = {}
        for row in deal_rows:
            cells = self.dao.list_cells_for_row(row.id)
            cells_by_row[row.id] = {c.column_id: c for c in cells}

        # Group rows by node_id, preserving order
        from collections import OrderedDict

        rows_by_node: OrderedDict[int, list[DealExportRow]] = OrderedDict()
        for row in deal_rows:
            rows_by_node.setdefault(row.node_id, []).append(row)

        # If no deal rows configured, fall back to one row per distribution step
        if not deal_rows:
            dist_steps = (
                self.db.query(ExecutionStep)
                .filter(ExecutionStep.run_id == run.id, ExecutionStep.node_type == "distribution")
                .order_by(ExecutionStep.step_order)
                .all()
            )
            out = io.StringIO()
            writer = csv.writer(out)
            writer.writerow([c.header_label for c in cols])
            for step in dist_steps:
                row_values = [self._resolve_column_default(c, run, deal, step) for c in cols]
                writer.writerow(row_values)
            return out.getvalue()

        # Generate with multi-row config
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow([c.header_label for c in cols])

        for node_id, node_rows in rows_by_node.items():
            for export_row in node_rows:
                row_cells = cells_by_row.get(export_row.id, {})
                csv_row = []
                for col in cols:
                    cell = row_cells.get(col.id)
                    if cell:
                        csv_row.append(self._resolve_cell(cell, context, run, deal, col))
                    else:
                        csv_row.append(self._resolve_column_default(col, run, deal, None))
                writer.writerow(csv_row)

        return out.getvalue()

    def _resolve_cell(
        self,
        cell: DealExportCell,
        context: dict[str, Decimal],
        run: ProcessingRun,
        deal: Deal | None,
        col: GlobalExportColumn,
    ) -> str:
        """Resolve a single cell value."""
        ref = cell.source_ref

        if cell.value_source == "node":
            val = context.get(ref, Decimal("0"))
            return self._format(val, col)

        if cell.value_source == "variable":
            val = context.get(ref, Decimal("0"))
            return self._format(val, col)

        if cell.value_source == "formula":
            try:
                val = self.formula_engine.execute(ref, context)
                return self._format(val, col)
            except Exception:
                return "ERROR"

        if cell.value_source == "literal":
            return ref

        if cell.value_source == "run_meta":
            return self._resolve_run_meta(ref, run)

        if cell.value_source == "deal_meta":
            return self._resolve_deal_meta(ref, deal)

        return ""

    def _resolve_column_default(
        self,
        col: GlobalExportColumn,
        run: ProcessingRun,
        deal: Deal | None,
        step: ExecutionStep | None,
    ) -> str:
        """Resolve a column using its template-level defaults (no DealExportCell)."""
        if col.value_type == "distribution_node":
            if step and step.result is not None:
                return self._format(step.result, col)
            return ""
        if col.value_type == "literal":
            return col.literal_value or ""
        if col.value_type == "run_meta":
            return self._resolve_run_meta(col.meta_field or "", run)
        if col.value_type == "deal_meta":
            return self._resolve_deal_meta(col.meta_field or "", deal)
        return ""

    @staticmethod
    def _resolve_run_meta(field: str, run: ProcessingRun) -> str:
        if field == "run_code":
            return f"RUN-{run.id}"
        if field in ("payment_date", "report_period"):
            return run.report_period or ""
        return ""

    @staticmethod
    def _resolve_deal_meta(field: str, deal: Deal | None) -> str:
        if not deal:
            return ""
        if field == "deal_id":
            return deal.name.replace(" ", "_") if deal.name else f"deal_{deal.id}"
        if field == "deal_name":
            return deal.name
        if field == "product_type":
            return deal.product_type or ""
        return ""

    @staticmethod
    def _format(value: Decimal | None, col: GlobalExportColumn) -> str:
        if value is None:
            return ""
        if col.format_type == "decimal" and isinstance(value, Decimal):
            places = col.decimal_places or 2
            return f"{value:.{places}f}"
        if col.format_type == "integer" and isinstance(value, Decimal):
            return f"{int(value)}"
        return str(value)

    @staticmethod
    def _placeholder(col: GlobalExportColumn) -> str:
        if col.value_type == "distribution_node":
            return "0.00"
        if col.value_type == "literal":
            return col.literal_value or ""
        if col.value_type == "run_meta":
            return f"<{col.meta_field or 'meta'}>"
        if col.value_type == "deal_meta":
            return f"<{col.meta_field or 'meta'}>"
        return ""
