"""Excel file reader utility."""

import openpyxl
from typing import Any


class ExcelReader:
    """Read servicer Excel tapes."""

    def __init__(self, file_path: str) -> None:
        self.wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()

    def get_sheet_names(self) -> list[str]:
        return self.wb.sheetnames

    def get_cell_value(self, sheet_name: str, column_letter: str, row_number: int) -> Any:
        ws = self.wb[sheet_name]
        cell_ref = f"{column_letter}{row_number}"
        return ws[cell_ref].value

    def get_sheet_grid(self, sheet_name: str, max_rows: int = 500) -> list[list[Any]]:
        ws = self.wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(list(row))
        return rows

    def read_sheet_grid(
        self,
        sheet_name: str,
        max_rows: int = 500,
        max_cols: int = 30,
    ) -> dict:
        """Return sheet contents as a navigable grid.

        Returns:
          {
            "sheet_name": str,
            "column_letters": ["A", "B", "C", ...],
            "rows": [
              {"row_number": 1, "cells": ["Wells Fargo", None, None, ...]},
              ...
            ],
          }
        """
        if sheet_name not in self.wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found.")

        ws = self.wb[sheet_name]
        rows_data: list[dict] = []

        actual_max_row = min(ws.max_row or 0, max_rows)
        actual_max_col = min(ws.max_column or 0, max_cols)

        column_letters = [self._col_num_to_letter(i) for i in range(1, actual_max_col + 1)]

        for row_idx in range(1, actual_max_row + 1):
            cells: list[Any] = []
            for col_idx in range(1, actual_max_col + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    cells.append(None)
                elif isinstance(value, (int, float)):
                    cells.append(value)
                else:
                    cells.append(str(value))
            rows_data.append({"row_number": row_idx, "cells": cells})

        return {
            "sheet_name": sheet_name,
            "column_letters": column_letters,
            "rows": rows_data,
        }

    @staticmethod
    def _col_num_to_letter(num: int) -> str:
        """Convert 1-indexed column number to Excel letter (1->A, 27->AA)."""
        result = ""
        while num > 0:
            num, rem = divmod(num - 1, 26)
            result = chr(65 + rem) + result
        return result

    def close(self) -> None:
        self.wb.close()
